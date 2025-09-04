import streamlit as st
import pandas as pd
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import plotly.express as px
import plotly.graph_objects as go

# Configuración de la página
st.set_page_config(
    page_title="TRR Monitoreo Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Códigos disponibles
CODIGOS_DISPONIBLES = [100, 102, 150, 151, 202, 203, 204, 205, 208, 210, 211, 231, 233, 244, 246, 251, 481]

def aplicar_colores_y_bordes_tabla(workbook, sheet_name='Resumen_Pivot'):
    """Aplica colores y bordes negros a la tabla"""
    try:
        ws = workbook[sheet_name]
        
        # Definir colores
        verde_claro = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        verde_oscuro = PatternFill(start_color='00F214', end_color='00F214', fill_type='solid')
        
        # Definir borde negro fino
        borde_negro = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Obtener dimensiones de la tabla
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 1. Aplicar a encabezados (fila 1) - verde oscuro + borde negro
        for col in range(1, max_col + 1):
            celda = ws.cell(row=1, column=col)
            celda.fill = verde_oscuro
            celda.border = borde_negro
        
        # 2. Aplicar a datos (filas 2 hasta penúltima) - verde claro + borde negro
        for row in range(2, max_row):
            for col in range(1, max_col + 1):
                celda = ws.cell(row=row, column=col)
                celda.fill = verde_claro
                celda.border = borde_negro
        
        # 3. Aplicar a fila de totales (última fila) - verde oscuro + borde negro
        for col in range(1, max_col + 1):
            celda = ws.cell(row=max_row, column=col)
            celda.fill = verde_oscuro
            celda.border = borde_negro
        
        return True
        
    except Exception as e:
        st.error(f"Error al aplicar colores y bordes: {str(e)}")
        return False

def procesar_archivo(df_completo, codigos_seleccionados):
    """
    Procesa el DataFrame según los códigos seleccionados
    """
    resultados = {}
    
    # Hoja 1: Datos sin filtrar
    resultados['datos_completos'] = df_completo.copy()
    
    # Crear hojas filtradas para cada código seleccionado
    if 'OverallReasonCode' in df_completo.columns:
        for codigo in codigos_seleccionados:
            # Convertir ambos a string para comparación más robusta
            mask = (df_completo['OverallReasonCode'].astype(str) == str(codigo))
            df_filtrado = df_completo[mask].copy()
            resultados[f'codigo_{codigo}'] = df_filtrado
    else:
        st.error("❌ No se encontró la columna 'OverallReasonCode' en el archivo")
        return None
    
    # Crear tabla dinámica - verificar que existe RequestID
    if 'RequestID' in df_completo.columns:
        pivot_data = df_completo.pivot_table(
            values='RequestID',
            index='OverallReasonCode',
            aggfunc='count'
        ).reset_index()
    else:
        # Si no existe RequestID, usar el índice para contar filas
        pivot_data = df_completo.groupby('OverallReasonCode').size().reset_index()
        pivot_data.columns = ['OverallReasonCode', 'Count']
    
    # Asegurar que las columnas tengan los nombres correctos
    if 'RequestID' in pivot_data.columns:
        pivot_data.columns = ['OverallReasonCode', 'Count']
    
    pivot_data['Percentage'] = (pivot_data['Count'] / pivot_data['Count'].sum()) * 100
    pivot_data['Percentage'] = pivot_data['Percentage'].round(2)
    
    # Agregar fila de totales
    total_row = pd.DataFrame({
        'OverallReasonCode': ['Total'],
        'Count': [pivot_data['Count'].sum()],
        'Percentage': [100.00]
    })
    pivot_data = pd.concat([pivot_data, total_row], ignore_index=True)
    
    resultados['pivot_table'] = pivot_data
    
    return resultados

def crear_excel_descarga(resultados, codigos_seleccionados):
    """
    Crea un archivo Excel en memoria para descargar
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir hoja de datos completos
        resultados['datos_completos'].to_excel(writer, sheet_name='Datos_Completos', index=False)
        
        # Escribir hojas filtradas
        for codigo in codigos_seleccionados:
            if f'codigo_{codigo}' in resultados:
                sheet_name = f'Codigo_{codigo}'
                resultados[f'codigo_{codigo}'].to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Escribir tabla dinámica
        resultados['pivot_table'].to_excel(writer, sheet_name='Resumen_Pivot', index=False)
    
    # Aplicar formato a la tabla dinámica
    workbook = load_workbook(output)
    aplicar_colores_y_bordes_tabla(workbook, 'Resumen_Pivot')
    
    # Guardar en nuevo buffer
    formatted_output = io.BytesIO()
    workbook.save(formatted_output)
    formatted_output.seek(0)
    
    return formatted_output.getvalue()

def crear_graficos(pivot_data):
    """
    Crea gráficos para visualizar los datos
    """
    # Filtrar datos sin el total
    data_sin_total = pivot_data[pivot_data['OverallReasonCode'] != 'Total'].copy()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Distribución por Código (Barras)")
        fig_bar = px.bar(
            data_sin_total, 
            x='OverallReasonCode', 
            y='Count',
            title='Cantidad de Registros por Código',
            color='Count',
            color_continuous_scale='Viridis'
        )
        fig_bar.update_layout(xaxis_title="Código", yaxis_title="Cantidad")
        st.plotly_chart(fig_bar, use_container_width=True)
    
    with col2:
        st.subheader("🥧 Distribución Porcentual (Torta)")
        fig_pie = px.pie(
            data_sin_total, 
            values='Count', 
            names='OverallReasonCode',
            title='Distribución Porcentual por Código'
        )
        st.plotly_chart(fig_pie, use_container_width=True)

def main():
    # Título principal
    st.title("📊 TRR Monitoreo Daily Analyzer")
    st.markdown("---")
    
    # Sidebar para configuración
    with st.sidebar:
        st.header("⚙️ Configuración")
        
        # Carga de archivo
        st.subheader("📁 Cargar Archivo")
        uploaded_file = st.file_uploader(
            "Selecciona tu archivo CSV",
            type=['csv'],
            help="Sube el archivo TRR_Monitoreo_Daily.csv"
        )
        
        # Selección de códigos
        st.subheader("🔍 Códigos a Filtrar")
        codigos_seleccionados = st.multiselect(
            "Selecciona los códigos a analizar:",
            options=CODIGOS_DISPONIBLES,
            default=[481],
            help="Puedes seleccionar múltiples códigos para crear hojas separadas"
        )
        
        # Opciones adicionales
        st.subheader("📋 Opciones")
        saltar_primera_fila = st.checkbox("Saltar primera fila", value=True)
        mostrar_graficos = st.checkbox("Mostrar gráficos", value=True)
    
    # Contenido principal
    if uploaded_file is not None:
        try:
            # Mostrar información del archivo
            st.success(f"✅ Archivo cargado: {uploaded_file.name}")
            
            # Cargar datos
            with st.spinner("📥 Cargando datos..."):
                skip_rows = 1 if saltar_primera_fila else 0
                df_completo = pd.read_csv(uploaded_file, sep=',', skiprows=skip_rows)
            
            st.info(f"📈 Datos cargados: {df_completo.shape[0]} filas, {df_completo.shape[1]} columnas")
            
            # Mostrar preview de los datos
            with st.expander("👀 Vista previa de los datos (primeras 5 filas)"):
                st.dataframe(df_completo.head())
            
            # Verificar que existe la columna necesaria
            if 'OverallReasonCode' not in df_completo.columns:
                st.error("❌ Error: No se encontró la columna 'OverallReasonCode' en el archivo")
                st.info("Columnas disponibles: " + ", ".join(df_completo.columns.tolist()))
                return
            
            # Procesar datos si hay códigos seleccionados
            if codigos_seleccionados:
                with st.spinner("⚙️ Procesando datos..."):
                    resultados = procesar_archivo(df_completo, codigos_seleccionados)
                
                if resultados:
                    # Mostrar resumen
                    st.subheader("📋 Resumen del Análisis")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Total de Registros", len(resultados['datos_completos']))
                    
                    with col2:
                        total_filtrados = sum([len(resultados[f'codigo_{codigo}']) for codigo in codigos_seleccionados if f'codigo_{codigo}' in resultados])
                        st.metric("Registros Filtrados", total_filtrados)
                    
                    with col3:
                        st.metric("Códigos Únicos", len(resultados['pivot_table']) - 1)  # -1 por el total
                    
                    # Mostrar detalles por código
                    st.subheader("🔍 Detalles por Código Seleccionado")
                    for codigo in codigos_seleccionados:
                        if f'codigo_{codigo}' in resultados:
                            cantidad = len(resultados[f'codigo_{codigo}'])
                            porcentaje = (cantidad / len(resultados['datos_completos'])) * 100
                            st.write(f"**Código {codigo}:** {cantidad} registros ({porcentaje:.2f}%)")
                    
                    # Tabla dinámica
                    st.subheader("📊 Tabla Dinámica - Resumen General")
                    st.dataframe(
                        resultados['pivot_table'], 
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Gráficos
                    if mostrar_graficos:
                        st.subheader("📈 Visualizaciones")
                        crear_graficos(resultados['pivot_table'])
                    
                    # Botón de descarga
                    st.subheader("💾 Descargar Análisis")
                    
                    with st.spinner("📦 Preparando archivo Excel..."):
                        excel_data = crear_excel_descarga(resultados, codigos_seleccionados)
                    
                    st.download_button(
                        label="📥 Descargar Análisis Completo (Excel)",
                        data=excel_data,
                        file_name=f"TRR_Monitoreo_Analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Descarga el archivo Excel con todas las hojas: datos completos, códigos filtrados y tabla dinámica"
                    )
                    
                    # Información adicional
                    with st.expander("ℹ️ Información del Archivo Excel"):
                        st.write("**El archivo Excel contiene las siguientes hojas:**")
                        st.write("- 📋 **Datos_Completos:** Todos los datos originales")
                        for codigo in codigos_seleccionados:
                            cantidad = len(resultados[f'codigo_{codigo}']) if f'codigo_{codigo}' in resultados else 0
                            st.write(f"- 🔍 **Codigo_{codigo}:** {cantidad} registros filtrados")
                        st.write("- 📊 **Resumen_Pivot:** Tabla dinámica con formato y colores")
            else:
                st.warning("⚠️ Selecciona al menos un código para procesar")
                
        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {str(e)}")
            st.exception(e)
    
    else:
        # Página de bienvenida
        st.markdown("""
        ## 👋 Bienvenido al Analizador de Monitoreo TRR
        
        Esta aplicación te permite analizar archivos CSV de monitoreo diario TRR con las siguientes funcionalidades:
        
        ### 🚀 Características principales:
        - 📁 **Carga de archivos:** Sube tu archivo CSV directamente
        - 🔍 **Filtrado personalizado:** Selecciona los códigos específicos que deseas analizar
        - 📊 **Tabla dinámica:** Visualiza un resumen completo con conteos y porcentajes
        - 📈 **Gráficos interactivos:** Visualizaciones en barras y torta
        - 💾 **Descarga Excel:** Obtén un archivo completo con formato profesional
        
        ### 📋 Códigos disponibles:
        `100, 102, 150, 151, 202, 203, 204, 205, 208, 210, 211, 231, 233, 244, 246, 251, 481`
        
        ### 🛠️ Para comenzar:
        1. Carga tu archivo CSV en la barra lateral
        2. Selecciona los códigos que deseas analizar
        3. Revisa los resultados y gráficos
        4. Descarga el análisis completo en Excel
        
        ---
        💡 **Tip:** Puedes seleccionar múltiples códigos para crear hojas separadas de cada uno en el archivo Excel final.
        """)

if __name__ == "__main__":
    main()